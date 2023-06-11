from django.shortcuts import render, redirect, HttpResponse
from studentManagement import models
from studentManagement.utils.pagination import Pagination

def depart_list(request):
    """部门列表"""

    # 去数据库中获取所有的部门列表
    queryset = models.Department.objects.all()

    page_object = Pagination(request, queryset)

    context = {
        'queryset': page_object.page_queryset,
        "page_string": page_object.html(),
    }

    return render(request, 'depart_list.html', context)

def depart_add(request):
    """添加部门"""
    if request.method == 'GET':
        return render(request, 'depart_add.html')

    # 获取用户POST过来的数据
    title = request.POST.get("title")

    # 保存到数据库
    models.Department.objects.create(title=title)

    # 重定向回部门列表
    return redirect("/depart/list")

def depart_delete(request):
    """删除部门"""
    # 获取id
    nid = request.GET.get("nid")

    # 删除
    models.Department.objects.filter(id=nid).delete()

    # 重定向回部门列表
    return redirect("/depart/list")

def depart_edit(request, nid):
    """修改部门"""

    # 根据nid，获取数据
    depart_name = models.Department.objects.filter(id=nid).first().title

    if request.method == "GET":
        return render(request, 'depart_edit.html', {'depart_name': depart_name})

    if request.method == "POST":
        # 获取新的部门名称
        new_title = request.POST.get("title")

        # 数据库更新
        models.Department.objects.filter(id=nid).update(title=new_title)

    # 重定向回部门列表
    return redirect("/depart/list")

def depart_multi(request):
    """批量上传（execl文件）"""
    from openpyxl import load_workbook

    # 1. 获取用户上传的文件对象
    file_object = request.FILES.get("exc")

    # 2. 对象传递给openpyxl，打开excel并读取内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]  #第0列的数据

    # 3. 循环获取每一段的数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    return redirect("/depart/list")
